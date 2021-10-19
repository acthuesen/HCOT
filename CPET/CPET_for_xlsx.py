import pandas as pd
from datetime import date, datetime, timedelta, time
from dateutil.relativedelta import relativedelta
import openpyxl as op
import os
import sys
sys._enablelegacywindowsfsencoding() #If your filepath includes special chars like æ,ø,å

###DIRECTORY AND FILENAMES####
d = 'YOURDIRECTORY'
fn = os.path.join(d, 'YOURFILE.xlsx'))

###CREATE TIMESERIES FOR MERGING###
#Define starts and stops and interval size
tstart = datetime(2000, 1, 1, 0, 0, 0) 
tstop = datetime(2000, 1, 1, 0, 30, 0) #this is 30 mins
delta = timedelta(seconds=1)
#Make list
times = []
while tstart < tstop: 
    times.append(tstart.strftime('%H:%M:%S'))
    tstart += delta
#Make df, put times in there, convert t to datetime dtype, set as index
tdf = pd.DataFrame()
tdf['t'] = times
tdf['tdt'] = pd.to_datetime(tdf['t'], format='%H:%M:%S')
tdf = tdf.set_index('tdt')

###IMPORT, CLEAN, AND AGGREGATE DATA###
full_data = pd.read_excel(fn)
#Extract ID and date, and height and weight, store as values
ID = full_data.columns[1]
date = full_data.columns[4]
wt = full_data.iloc[5,1]
ht = full_data.iloc[4,1]
#drop cols 0-8/select cols between t and PaCO2_e (pii and test parms df); drop rows 1-2
subset_data = full_data.loc[:, 't':'PaCO2_e']
vdat = subset_data.drop([0,1])
#Infer dtypes
vdat = vdat.infer_objects()
#convert t to datetime dtype stored as tdt (to be used as index)
vdat['tdt'] = pd.to_datetime(vdat['t'], format='%H:%M:%S')
#aggregate data by calculating means for num vars, firsts for cat vars
agg_vdat = vdat.groupby('tdt').agg(
	{
		'VO2':'mean',
		'VCO2':'mean',
		'RQ':'mean',
		't':'first',
		'ID':'first',
		'date':'first',
		'filename':'first',
		'HR':'max'
		}
	)
#Concatenate data (by index)
justtime = agg_vdat[['t']]
final_time = justtime.iloc[-1,0]
split_final_time = final_time.split(":")
final_hrs = int(split_final_time[0])
final_mns = int(split_final_time[1])
final_scs = int(split_final_time[2])
final_hrs_is = final_hrs*60*60
final_mns_is = final_mns*60
final_scs_is = final_scs
final_time_is = final_hrs_is + final_mns_is + final_scs_is
ntdf = tdf[0:final_time_is+1]
eagg_vdat = pd.concat([ntdf,agg_vdat],axis=1)
#calculate rolling averages of VO2, VCO2 and RQ. 
eagg_vdat['VO2_20s'] = eagg_vdat['VO2'].rolling(20, center=False, min_periods=1).mean()
eagg_vdat['VCO2_20s'] = eagg_vdat['VCO2'].rolling(20, center=False, min_periods=1).mean()
eagg_vdat['RQ_20s'] = eagg_vdat['RQ_20s'].rolling(20, center=False, min_periods=1).mean()
#Find max's calc bmi and ftnlvl
VO2_20s_max = eagg_vdat['VO2_20s'].max()
VCO2_20s_max = eagg_vdat['VCO2_20s'].max()
RQ_20s_max = eagg_vdat['RQ'].max()
HR_max = eagg_vdat['HR'].max()
ftnlvl = VO2_20s_max/wt
bmi = wt/(ht/100)**2

###ADD CALCULATED VALUES TO RESULTS SHEET###
rfn = os.path.join(d, 'YOURRESULTFILE.xlsx')
wb = op.load_workbook(rfn)
ws = wb.active
maxr = ws.max_row + 1
ws.cell(row = maxr, column = 1).value = ID
ws.cell(row = maxr, column = 2).value = ht
ws.cell(row = maxr, column = 3).value = wt
ws.cell(row = maxr, column = 4).value = bmi
ws.cell(row = maxr, column = 5).value = VO2_20s_max
ws.cell(row = maxr, column = 6).value = VCO2_20s_max
ws.cell(row = maxr, column = 7).value = RQ_20s_max
ws.cell(row = maxr, column = 8).value = ftnlvl
ws.cell(row = maxr, column = 9).value = HRmax
ws.cell(row = maxr, column = 10).value = date
ws.cell(row = maxr, column = 11).value = filename
wb.save(rfn)
