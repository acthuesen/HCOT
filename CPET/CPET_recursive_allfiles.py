#Depending on setup, it's useful to do this for many files at a time. 
#This works for xls and xlsx files and is set up to move files from a folder within the directory called 'new' to a folder within the directory called 'calc'. I.e. during data collection new files are placed in 'new' folder and calculated as required. 
#Results are added to a results file which can be used for downstream statistical procedures eg. 

import pandas as pd
from datetime import date, datetime, timedelta, time
from dateutil.relativedelta import relativedelta
import openpyxl as op
import os
import glob
import shutil
import xlrd
import csv
import sys
sys.__stdout__ = sys.stdout #This fixes the fileno error you get when importing a file
sys._enablelegacywindowsfsencoding() #This fixes the problem with non ascii chr in dir

###CREATE TIMESERIES FOR MERGING###
#Define starts and stops and interval size
tstart = datetime(2000, 1, 1, 0, 0, 0) 
tstop = datetime(2000, 1, 1, 0, 30, 0) 
delta = timedelta(seconds=1)
#Make list
times = []
while tstart < tstop: 
    times.append(tstart.strftime('%H:%M:%S'))
    tstart += delta
#Make df, put times in there, convert t to datetime dtype, set as index
tdf = pd.DataFrame()
tdf['t'] = times
tdf['tdt'] = pd.to_datetime(tdf['t'], format='%H:%M:%S')
tdf = tdf.set_index('tdt')

###DIRECTORY AND FILENAMES####
#Make list of files
d = 'YOURPATH'
filenames = glob.glob(os.path.join(d, '*new\\', '*.xls'))
filenamesx = glob.glob(os.path.join(d, '*new\\', '*.xlsx'))

###CALCULATIONS####
#Split for xls (old COSMED app) and xlsx files (new COSMED app) because the sheets are setup differently, and because of the unicode error the old xls files give
										
###FOR XLSX FILES###
for filename in filenamesx:
###IMPORT, CLEAN, AND AGGREGATE DATA###
	full_data = pd.read_excel(filename)
#Extract ID and date, and height and weight, store as values
	ID = full_data.columns[1]
	date = full_data.columns[4]
	wt = full_data.iloc[5,1]
	ht = full_data.iloc[4,1]
#drop cols 0-8/select cols between t and PaCO2_e; drop rows 1-2
	subset_data = full_data.loc[:, 't':'PaCO2_e']
	vdat = subset_data.drop([0,1])
#Infer dtypes
	vdat = vdat.infer_objects()
#convert t to datetime dtype stored as tdt (to be used as index)
	vdat['tdt'] = pd.to_datetime(vdat['t'], format='%H:%M:%S')
#Create new columns with ID, date, and filename
	vdat['ID'] = ID
	vdat['date'] = date
	vdat['filename'] = filename
	vdat['weight'] = wt
	vdat['height'] = ht
#aggregate data by calculating means, firsts for cat vars
	agg_vdat = vdat.groupby('tdt').agg(
		{
			'VO2':'mean',
			'VCO2':'mean',
			'RQ':'mean',
			't':'first',
			'ID':'first',
			'date':'first',
			'filename':'first',
			'HR':'max'
			}
		)
#Concatenate data (by index)
	justtime = agg_vdat[['t']]
	final_time = justtime.iloc[-1,0]
	split_final_time = final_time.split(":")
	final_hrs = int(split_final_time[0])
	final_mns = int(split_final_time[1])
	final_scs = int(split_final_time[2])
	final_hrs_is = final_hrs*60*60
	final_mns_is = final_mns*60
	final_scs_is = final_scs
	final_time_is = final_hrs_is + final_mns_is + final_scs_is
	ntdf = tdf[0:final_time_is+1]
	eagg_vdat = pd.concat([ntdf,agg_vdat],axis=1)
#calculate rolling averages of VO2, VCO2 and RQ. 
	eagg_vdat['VO2_20s'] = eagg_vdat['VO2'].rolling(20, center=False, min_periods=1).mean()
	eagg_vdat['VCO2_20s'] = eagg_vdat['VCO2'].rolling(20, center=False, min_periods=1).mean()
	eagg_vdat['RQ_20s'] = eagg_vdat['RQ_20s'].rolling(20, center=False, min_periods=1).mean()
#Find max's calc bmi and ftnlvl
	VO2_20s_max = eagg_vdat['VO2_20s'].max()
	VCO2_20s_max = eagg_vdat['VCO2_20s'].max()
	RQ_20s_max = eagg_vdat['RQ'].max()
	HR_max = eagg_vdat['HR'].max()
	ftnlvl = VO2_20s_max/wt
	bmi = wt/(ht/100)**2
###ADD CALCULATED VALUES TO RESULTS SHEET###
	rfn = os.path.join(d, 'vo2max_results.xlsx')
	wb = op.load_workbook(rfn)
	ws = wb.active
	maxr = ws.max_row + 1
	ws.cell(row = maxr, column = 1).value = ID
	ws.cell(row = maxr, column = 2).value = ht
	ws.cell(row = maxr, column = 3).value = wt
	ws.cell(row = maxr, column = 4).value = bmi
	ws.cell(row = maxr, column = 5).value = VO2_20s_max
	ws.cell(row = maxr, column = 6).value = VCO2_20s_max
	ws.cell(row = maxr, column = 7).value = RQ_20s_max
	ws.cell(row = maxr, column = 8).value = ftnlvl
	ws.cell(row = maxr, column = 9).value = HRmax
	ws.cell(row = maxr, column = 10).value = date
	ws.cell(row = maxr, column = 11).value = filename
	wb.save(rfn)
	

###XLS FILES###
###Start for loop###
for filename in filenames:
###IMPORT, CLEAN, AND AGGREGATE DATA###
	fd = xlrd.open_workbook(filename, encoding_override='1252')
	sh = fd.sheet_by_index(0)
	csv_fd = open(os.path.join(d, 'new\\csv_fd.csv'), 'w', newline='')
	wr = csv.writer(csv_fd, quoting=csv.QUOTE_ALL)
	for rownum in range(sh.nrows):
		wr.writerow(sh.row_values(rownum))
	csv_fd.close()
	full_data = pd.read_csv(os.path.join(d, 'new\\csv_fd.csv'), encoding="1252")
#Extract ID and date, and height and weight, store as values
#Nb, because everything is an object, remember to store wt and ht as floats
	ID = full_data.columns[1]
	date = full_data.iloc[0,4]
	wt = float(full_data.iloc[5,1])
	ht = float(full_data.iloc[4,1])
#drop cols 0-8/subset from t to User 3, drop rows 1-2
	subset_data = full_data.loc[:, 't':'User 3']
	vdat = subset_data.drop([0,1])
#Infer dtypes (infer datatypes doesn't work on this file like it does for xlsx files
	vdat[['VO2','VCO2','R','HR']] = vdat[['VO2','VCO2','R','HR']].apply(pd.to_numeric)
#convert t to datetime dtype stored as tdt (to be used as index)
	vdat['tdt'] = pd.to_datetime(vdat['t'], format='%H:%M:%S')
#Create new columns with ID, date, and filename
	vdat['ID'] = ID
	vdat['date'] = date
	vdat['filename'] = filename
	vdat['weight'] = wt
	vdat['height'] = ht
#aggregate data by calculating means, firsts for cat vars
	agg_vdat = vdat.groupby('tdt').agg(
		{
			'VO2':'mean',
			'VCO2':'mean',
			'R':'mean',
			't':'first',
			'ID':'first',
			'date':'first',
			'filename':'first',
			'HR':'max'
			}
		)
#Concatenate data (by index)
	justtime = agg_vdat[['t']]
	final_time = justtime.iloc[-1,0]
	split_final_time = final_time.split(":")
	final_hrs = int(split_final_time[0])
	final_mns = int(split_final_time[1])
	final_scs = int(split_final_time[2])
	final_hrs_is = final_hrs*60*60
	final_mns_is = final_mns*60
	final_scs_is = final_scs
	final_time_is = final_hrs_is + final_mns_is + final_scs_is
	ntdf = tdf[0:final_time_is+1]
	eagg_vdat = pd.concat([ntdf,agg_vdat],axis=1)
#calculate rolling averages. Obs min_periods means that it calculates on the basis of one value if there are NaNs
	eagg_vdat['VO2_20s'] = eagg_vdat['VO2'].rolling(20, center=False, min_periods=1).mean()
	eagg_vdat['VCO2_20s'] = eagg_vdat['VCO2'].rolling(20, center=False, min_periods=1).mean()
	eagg_vdat['R_20s'] = eagg_vdat['R'].rolling(20, center=False, min_periods=1).mean()
#Find max's, bmi and ftnlvl
	VO2_20s_max = eagg_vdat['VO2_20s'].max()
	VCO2_20s_max = eagg_vdat['VCO2_20s'].max()
	R_20s_max = eagg_vdat['R_20s'].max()
	HR_max = eagg_vdat['HR'].max()
	ftnlvl = VO2_20s_max/wt
	bmi = wt/(ht/100)**2
###ADD CALCULATED VALUES TO RESULTS SHEET### 
	rfn = os.path.join(d, 'vo2max_results.xlsx')
	wb = op.load_workbook(rfn)
	ws = wb.active
	maxr = ws.max_row + 1
	ws.cell(row = maxr, column = 1).value = ID
	ws.cell(row = maxr, column = 2).value = ht
	ws.cell(row = maxr, column = 3).value = wt
	ws.cell(row = maxr, column = 4).value = bmi
	ws.cell(row = maxr, column = 5).value = VO2_20s_max
	ws.cell(row = maxr, column = 6).value = VCO2_20s_max
	ws.cell(row = maxr, column = 7).value = R_20s_max
	ws.cell(row = maxr, column = 8).value = ftnlvl
	ws.cell(row = maxr, column = 9).value = HR_max
	ws.cell(row = maxr, column = 10).value = date
	ws.cell(row = maxr, column = 11).value = filename
	wb.save(rfn)
###Delete csv file
	os.remove(os.path.join(d, 'new\\csv_fd.csv'))

af = glob.glob(os.path.join(d, '*new\\*'))
source = os.path.join(d, 'new')
dest = os.path.join(d, 'calc')

for f in af:
	shutil.move(f,dest)
